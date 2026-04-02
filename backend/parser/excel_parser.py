"""
Excel parser — reads the Activity Capture Template workbook.

Each employee submits their OWN copy of the template.
Only the "Survey" sheet contains real data — all "Smpl-*" sheets are
sample/guidance examples embedded in the template and must be ignored.

Sheet layout (verified against live file):
  Row 4  col B → Employee Name
  Row 5  col B → Employee Number
  Row 6  col B → Function
  Row 7  col B → Job Title
  Row 8  col B → Direct Manager

  First row where col A == "#"  → table header sentinel
  Rows below: col A is an integer row index; stop when it is not.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

import openpyxl


SURVEY_SHEET = "Survey"  # the only sheet with real employee data

HEADER_MAP = {  # row number → employee field name
    4: "employee_name",
    5: "employee_number",
    6: "function",
    7: "job_title",
    8: "direct_manager",
}

TABLE_COLS = {
    # col index (1-based) → field name
    1: "row_index",
    2: "description",
    3: "frequency",
    4: "pct_time",
    5: "output_deliverable",
    6: "who_uses_it",
    7: "value_type",
    8: "if_stopped",
    9: "can_be_automated",
}


@dataclass
class ParsedActivity:
    row_index: int
    description: str
    frequency: str | None = None
    pct_time: float | None = None
    output_deliverable: str | None = None
    who_uses_it: str | None = None
    value_type: str | None = None
    if_stopped: str | None = None
    can_be_automated: str | None = None


@dataclass
class ParsedEmployee:
    sheet_name: str
    employee_name: str | None = None
    employee_number: str | None = None
    function: str | None = None
    job_title: str | None = None
    direct_manager: str | None = None
    activities: list[ParsedActivity] = field(default_factory=list)
    source_filename: str | None = None   # set by upload router for traceability


def _cell_str(ws: Any, row: int, col: int) -> str | None:
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    return str(val).strip() or None


def _parse_sheet(ws: Any) -> ParsedEmployee:
    emp = ParsedEmployee(sheet_name=ws.title)

    # Header block — rows 4-8, value always in col B (col index 2)
    for row_num, attr in HEADER_MAP.items():
        setattr(emp, attr, _cell_str(ws, row_num, 2))

    # Find the sentinel row where col A == "#"
    sentinel_row: int | None = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if str(row[0].value).strip() == "#":
            sentinel_row = row[0].row
            break

    if sentinel_row is None:
        return emp  # no activity table found

    # Parse activity rows below sentinel
    for row in ws.iter_rows(min_row=sentinel_row + 1, max_row=ws.max_row):
        raw_idx = row[0].value
        # Stop when col A is no longer an integer
        try:
            idx = int(raw_idx)
        except (TypeError, ValueError):
            break

        desc = str(row[1].value).strip() if row[1].value else ""
        if not desc:
            break  # blank description = end of data

        pct_raw = row[3].value
        try:
            pct = float(pct_raw) if pct_raw is not None else None
        except (TypeError, ValueError):
            pct = None

        activity = ParsedActivity(
            row_index=idx,
            description=desc,
            frequency=str(row[2].value).strip() if row[2].value else None,
            pct_time=pct,
            output_deliverable=str(row[4].value).strip() if row[4].value else None,
            who_uses_it=str(row[5].value).strip() if row[5].value else None,
            value_type=str(row[6].value).strip() if row[6].value else None,
            if_stopped=str(row[7].value).strip() if row[7].value else None,
            can_be_automated=str(row[8].value).strip() if row[8].value else None,
        )
        emp.activities.append(activity)

    return emp


def parse_workbook(file_bytes: bytes) -> ParsedEmployee | None:
    """
    Parse an employee's submitted xlsx workbook from raw bytes.
    Reads ONLY the 'Survey' sheet — all Smpl-* sheets are template examples, not data.
    Returns a ParsedEmployee, or None if the Survey sheet is missing or empty.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), read_only=True, data_only=True
    )
    if SURVEY_SHEET not in wb.sheetnames:
        wb.close()
        return None

    ws = wb[SURVEY_SHEET]
    emp = _parse_sheet(ws)
    wb.close()

    if not emp.activities:
        return None

    return emp
